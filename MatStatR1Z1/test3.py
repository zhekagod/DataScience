import openpyxl
import numpy as np
import math as m
import matplotlib.pyplot as plt
import seaborn
import random

# num = np.nanmedian
# print(int(input()))

book = openpyxl.open("my_r1z1.xlsx", read_only=True)

sheet = book.active

print(sheet["A3"].value)

print(sheet[2][0].value)

# selection = [sheet[f"A{row}"].value for row in range(1, sheet.max_row)]
# Нумерация рядов с 1, а колонок с 0
selection = [sheet[row][0].value for row in range(2, sheet.max_row + 1)]

selection = sorted(selection)

cells = sheet['A2':'A88']

'''for num in cells:
    print(num)'''

maxim_row = sheet.max_row
maxim_column = sheet.max_column
print(maxim_row)
stroka = [sheet[2][0].value for i in range(1, sheet.max_row)]

print(*selection)

select_size = len(selection)  # объём выборки
min_select = min(selection)  # минимум
max_select = max(selection)  # максимум
scale = max_select - min_select  # размах
counting = int(select_size / 10)
width = scale / counting  # толщина столбцов


# (интервал) дельта
# print(f"widht = {width}")


def expected_value(sel, sel_size):  # мат.ожидание
    return sum(sel) / sel_size


print("============================================================")
print(f"Объём выборки: {select_size}")
print(f"Минимум: {min_select}")
print(f"Максимум: {max_select}")
print(f"Размах: {scale}")
print(f"Мат ожидание: "
      f"{expected_value(selection, select_size)}")


def sq(x):
    return x ** 2


def variance(sel, sel_size):  # смещённая дисперсия
    result = 0
    exp = expected_value(sel, sel_size)
    for i in range(len(sel)):
        result += sq(sel[i] - exp)
    return result / sel_size


print(f"Смещённая выборочная дисперсия: "
      f"{variance(selection, select_size)}")


def unbiased_variance(sel, sel_size):  # несмещённая дисперсия
    return (variance(sel, sel_size) * sel_size) / (sel_size - 1)


print(f"Несмещённая выборочная дисперсия: "
      f"{unbiased_variance(selection, select_size)}")


def deviation(sel, sel_size):
    return m.sqrt(variance(sel, sel_size))


print(f"Стандартное отклонение: {deviation(selection, select_size)}")


def assymerty(sel, sel_size):
    result = 0
    exp = expected_value(sel, sel_size)
    for i in range(len(sel)):
        result += (sel[i] - exp) ** 3
    return result / (sel_size * (deviation(sel, sel_size) ** 3))


print(f"Ассиметрия: {assymerty(selection, select_size)}")


def median(sel, sel_size):
    if sel_size % 2 == 1:
        return sel[int(sel_size / 2)]
    return (sel[int(sel_size / 2)] + sel[int(sel_size / 2) - 1]) / 2


print("Медиана: ", median(selection, select_size))


def shirota(sel, sel_size, q):
    if (sel_size - 1) * q == int((sel_size - 1) * q):
        return sel[(sel_size - 1) * q + 1]
    elif (sel_size - 1) * q > int((sel_size - 1) * q):
        return (sel[int((sel_size - 1) * q) + 1] + sel[int((sel_size - 1) * q) + 2]) / 2


print("Интерквартильная широта: ", shirota(selection, select_size, (3 / 4)) -
      shirota(selection, select_size, (1 / 4)))
print("============================================================")

fig = plt.figure(figsize=(6, 4))
ax = fig.add_subplot()



"""x_list = np.arange(min_select, max_select + width - 0.1, width)
x1 = np.linspace(min(selection), max(selection), int(select_size / 10))


def create_y_column(x, sel, sel_size):
    arr = []
    for i in range(1, len(x_list)):
        count = 0
        for num in selection:
            if x_list[i - 1] <= num <= x_list[i]:
                count += 1
        arr.append(count / (select_size * (x_list[i] - x_list[i - 1])))
    return arr


x = np.linspace(min_select, max_select, counting)
list_y = [0 for i in range(len(x1) - 1)]

list_x = list()
for i in range(len(x) - 1):
    if i < (len(x) - 1) / 2:
        list_x.append(x[i] + 0.75)
    else:
        list_x.append(x[i] + 0.40)
y1 = create_y_column(x, selection, select_size)"""


# x = np.linspace(min_select, max_select, int(m.log(select_size, 2)))
# x = np.linspace(min_select, max_select, counting + 1)
x = np.linspace(min_select, max_select, counting + 5)
bar_y = [0 for i in range(len(x) - 1)]


# Подсчёт высоты столбца
def create_y_column(x, sel, sel_size):
    arr = []
    for i in range(1, len(x)):
        count = 0
        for num in selection:
            if x[i - 1] <= num <= x[i]:
                count += 1
        arr.append(count / (select_size * (x[i] - x[i - 1])))
        # Так как гистограмма вероятностная, то делим ещё на дельту
    return arr


y1 = create_y_column(x, selection, select_size)
"""for i in range(1, len(x)):
    for j in range(ind, len(selection)):
        if x[i - 1] <= selection[j] < x[i]:
            bar_y[i - 1] += 1
            ind += 1
bar_y = [i / (x[1] - x[0]) / select_size for i in bar_y]"""


list_x = list()
#Костыльное выпраление ширины столбцов
for i in range(len(x) - 1):
    if i < (len(x) - 1) / 2:
        list_x.append(x[i] + 0.75) # + 0.75
    else:
        list_x.append(x[i] + 0.40) # + 0.4

# Вывод гистограммы
ax.bar(list_x, y1, width=width)
# Вывод функции плотности
seaborn.kdeplot(selection, color="red")
ax.grid()

plt.show()


# Построение и вывод ЭФР
def empirical_dist_func(sel, sel_size):
    count = 1
    for i in range(sel_size - 1):
        if not (sel[i] == sel[i + 1]):
            plt.plot([sel[i], sel[i + 1]], [count / sel_size, count / sel_size])
        count += 1
    plt.plot([sel[sel_size - 1], sel[sel_size - 1] + 1], [1, 1])
    plt.show()


empirical_dist_func(selection, select_size)
