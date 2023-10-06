import openpyxl
import numpy as np
import math as m
import matplotlib.pyplot as plt
import random

import seaborn
from scipy.stats import norm

# num = np.nanmedian
# print(int(input()))

book = openpyxl.open("my_r1z1.xlsx", read_only=True)

sheet = book.active

print(sheet["A3"].value)

print(sheet[2][0].value)

# selection = [sheet[f"A{row}"].value for row in range(1, sheet.max_row)]

selection = [sheet[row][0].value for row in range(2, sheet.max_row + 1)]

selection = sorted(selection)

cells = sheet['A2':'A88']

'''for num in cells:
    print(num)'''

maxim_row = sheet.max_row
maxim_column = sheet.max_column
print(maxim_row)
stroka = [sheet[2][0].value for i in range(1, sheet.max_row)]

# print(*stroka)

print(*selection)

select_size = len(selection)  # объём выборки
min_select = min(selection)  # минимум
max_select = max(selection)  # максимум
print(f"min = {min_select}" + f", max = {max_select}")
scale = max_select - min_select  # размах
exp_value = sum(selection) / select_size
exp_value = 121.3391
# scale / int(n/10)
width = scale / (int(select_size / 10))  # толщина столбцов
# (интервал) дельта
print(f"widht = {width}")


def expected_value(sel, sel_size):  # мат.ожидание
    return sum(sel) / sel_size


print("============================================================")
print(f"Объём выборки: {select_size}")
print(f"Минимум: {min_select}")
print(f"Максимум: {max_select}")
print(f"Размах: {scale}")
print(f"Мат ожидание: "
      f"{expected_value(selection, select_size)}")


def sq(x):
    return x ** 2


def variance(sel, sel_size):  # смещённая дисперсия
    result = 0
    exp = expected_value(sel, sel_size)
    for i in range(len(sel)):
        result += sq(sel[i] - exp)
    return result / sel_size


print(f"Смещённая выборочная дисперсия: "
      f"{variance(selection, select_size)}")


def unbiased_variance(sel, sel_size):  # несмещённая дисперсия
    return (variance(sel, sel_size) * sel_size) / (sel_size - 1)


print(f"Несмещённая выборочная дисперсия: "
      f"{unbiased_variance(selection, select_size)}")


def deviation(sel, sel_size):
    return m.sqrt(variance(sel, sel_size))


print(f"Стандартное отклонение: {deviation(selection, select_size)}")


def assymerty(sel, sel_size):
    result = 0
    exp = expected_value(sel, sel_size)
    for i in range(len(sel)):
        result += (sel[i] - exp) ** 3
    return result / (sel_size * (deviation(sel, sel_size) ** 3))


print(f"Ассиметрия: {assymerty(selection, select_size)}")


def median(sel, sel_size):
    if sel_size % 2 == 1:
        return sel[int(sel_size / 2)]
    return (sel[int(sel_size / 2)] + sel[int(sel_size / 2) - 1]) / 2


print("Медиана: ", median(selection, select_size))


def shirota(sel, sel_size, q):
    if (sel_size - 1) * q == int((sel_size - 1) * q):
        return sel[(sel_size - 1) * q + 1]
    elif (sel_size - 1) * q > int((sel_size - 1) * q):
        return (sel[int((sel_size - 1) * q) + 1] + sel[int((sel_size - 1) * q) + 2]) / 2


print("Интерквартильная широта: ", shirota(selection, select_size, (3 / 4)) -
      shirota(selection, select_size, (1 / 4)))
print("============================================================")

print(expected_value(selection, select_size))

x_list = np.arange(min_select, max_select + width - 0.1, width)


# print(*x_list)


def correct_x_list(nums):
    for i in range(len(nums)):
        if nums[i] - int(nums[i]) == 0:
            nums[i] = int(nums[i])


var = variance(selection, select_size)
dev = deviation(selection, select_size)
correct_x_list(x_list)
print(x_list)
print(f"variance = {var}")
print(f"devitation = {dev}")
dev = round(dev, 6)
exp_value = round(exp_value, 6)
print(dev)

# otrezki =
# y1_list = [len([item for i, item in selection if x_list[i] <= item <= x_list[i+1]])]
y1_list = []
for i in range(len(x_list) - 1):
    count = 0
    for num in selection:
        if x_list[i] <= num <= x_list[i + 1]:
            count += 1
    y1_list.append(count / (select_size * width))
print(*y1_list)
print(len(y1_list))
"""x_indexes = np.arange(len(x_list))
plt.bar(x_indexes - (width/2), y1_list, label="Diagram",
        width=width)
plt.show()"""

"""fig, ax = plt.subplots()
x = np.linspace(-5, 5, 100)
y = -x ** 2
ax.plot(x, y)
plt.show()"""

fig = plt.figure(figsize=(6, 4))
ax = fig.add_subplot()
x = np.linspace(-5, 5, 87)
x1 = np.linspace(min(selection), max(selection), int(select_size/10))

y1 = []


def create_y_column(x, sel, sel_size):
    arr = []
    for i in range(1, len(x_list)):
        count = 0
        for num in selection:
            if x_list[i-1] <= num <= x_list[i]:
                count += 1
        arr.append(count / (select_size * (x_list[i]-x_list[i-1])))
    return arr


y1 = create_y_column(x1, selection, select_size - 1)
listx = []
listx = [x1[i] + 0.75 if i < (len(x1)) / 2
         else x1[i]+0.4 for i in range(len(x1))]

"""for i in range(len(x1)):
    if i < (len(x1)) / 2:
        listx.append(x1[i] + 0.7)
    else:
        listx.append(x1[i] + 0.4)"""
#=============================================
height = []

#=========================================

print(f"listx = {listx}")
ax.bar(listx, y1, width=width)
seaborn.kdeplot(selection, color="red")
ax.grid()
plt.show()

# y = (1 / (m.sqrt(2) * m.sqrt(m.pi * dev))) * m.exp(-0.5 * ((sq(x - exp_value)) / var))  # вручную перебрать


# и написать
# y = 1/x1


def func(x):
    return (1 / (m.sqrt(2) * m.sqrt(m.pi * dev))) \
           * m.exp(-0.5 * ((sq(x - exp_value)) / var))


y = [func(num) for num in selection]

print(*y)

x_values = np.arange(0, width * int(select_size / 10), width)

# ax.plot(x_values, y)


"""ax.hist(y)
ax.grid()"""

y_column = random.sample(selection,
                         int(select_size / 10))
x_len = x_values
print(*y_column)
print("Другая гистограмма: ")
plt.bar(x_len, y_column)  # ок
"""x1 = np.arange(-5, 5, 0.001)
plt.plot(x1, norm.pdf(x1, 0, 1))
plt.show()"""

"""x1 = np.arange(min(y1_list), max(y1_list), 0.001)
plt.plot(x1, norm.pdf(x1, exp_value, dev))
plt.show()"""
"""plt.gca().invert_xaxis()
plt.gca().invert_yaxis()"""
#plt.plot(selection, norm.pdf(selection, exp_value, dev))
# plt.plot(norm.pdf(selection, exp_value, var), selection)
# plt.ylim(0, 130)
plt.show()

"""x1 = np.arange(-5, 5, 0.001)
plt.plot(x1, norm.pdf(x1, exp_value, dev))
plt.show()"""

'''for row in range(1, 10):
    value = sheet[row][0].value
    print(value)'''
